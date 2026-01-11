# Execution Verifier
"""
Executes code in a Docker sandbox and verifies outputs.

Key principles:
- NEUTRAL ENVIRONMENT: python:3.10-slim with NO pre-installed packages
- Model must install its own dependencies (tests if skill teaches complete workflow)
- 60 second timeout, 512MB memory limit
- Content verification (not just existence)
"""

import json
import os
import shutil
import subprocess
import tempfile
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from evaluator.code_extractor import extract_code, ExtractedCode
from evaluator.skill_categories import SkillCategory, get_skill_category


@dataclass
class ExecutionResult:
    """Result of executing code in the sandbox."""
    executed: bool
    exit_code: int
    stdout: str
    stderr: str
    execution_time_ms: int
    output_files: list[str]
    error: Optional[str] = None


@dataclass
class VerificationResult:
    """Result of verifying execution output."""
    skill_name: str
    prompt: str

    # Code extraction
    code_extracted: bool
    code_language: str
    code_blocks_count: int

    # Execution
    executed: bool
    execution_success: bool
    execution_error: Optional[str]
    execution_time_ms: int

    # Output validation
    output_files_created: list[str]
    output_valid: bool
    output_properties: dict = field(default_factory=dict)

    # Metadata
    verified_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))


# Docker image for execution - NEUTRAL, no pre-installed packages
DOCKER_IMAGE = "python:3.10-slim"
EXECUTION_TIMEOUT_SECONDS = 60
MEMORY_LIMIT = "512m"


def is_docker_available() -> bool:
    """Check if Docker is available on the system."""
    try:
        result = subprocess.run(
            ["docker", "version"],
            capture_output=True,
            timeout=5,
        )
        return result.returncode == 0
    except (subprocess.TimeoutExpired, FileNotFoundError):
        return False


def execute_in_docker(
    code: str,
    language: str,
    work_dir: Path,
    timeout: int = EXECUTION_TIMEOUT_SECONDS,
) -> ExecutionResult:
    """
    Execute code in a Docker container.

    Args:
        code: The code to execute
        language: Programming language (python, bash, etc.)
        work_dir: Working directory (mounted into container)
        timeout: Execution timeout in seconds

    Returns:
        ExecutionResult with exit code, output, and timing
    """
    start_time = datetime.now()

    # Write code to file
    if language == "python":
        script_file = work_dir / "script.py"
        script_file.write_text(code)
        cmd_in_container = "python script.py"
    elif language == "bash":
        script_file = work_dir / "script.sh"
        script_file.write_text(code)
        cmd_in_container = "bash script.sh"
    else:
        return ExecutionResult(
            executed=False,
            exit_code=-1,
            stdout="",
            stderr=f"Unsupported language: {language}",
            execution_time_ms=0,
            output_files=[],
            error=f"Unsupported language: {language}",
        )

    # Build docker command
    docker_cmd = [
        "docker", "run",
        "--rm",  # Remove container after execution
        "--network", "host",  # Allow network for pip install
        "--memory", MEMORY_LIMIT,
        "--cpus", "1",
        "-v", f"{work_dir.absolute()}:/workspace",
        "-w", "/workspace",
        DOCKER_IMAGE,
        "sh", "-c", cmd_in_container,
    ]

    try:
        result = subprocess.run(
            docker_cmd,
            capture_output=True,
            timeout=timeout,
            text=True,
        )

        execution_time_ms = int((datetime.now() - start_time).total_seconds() * 1000)

        # List output files (excluding the script itself)
        output_files = [
            f.name for f in work_dir.iterdir()
            if f.is_file() and f.name not in ("script.py", "script.sh")
        ]

        return ExecutionResult(
            executed=True,
            exit_code=result.returncode,
            stdout=result.stdout,
            stderr=result.stderr,
            execution_time_ms=execution_time_ms,
            output_files=output_files,
        )

    except subprocess.TimeoutExpired:
        execution_time_ms = int((datetime.now() - start_time).total_seconds() * 1000)
        return ExecutionResult(
            executed=True,
            exit_code=-1,
            stdout="",
            stderr="Execution timed out",
            execution_time_ms=execution_time_ms,
            output_files=[],
            error="Timeout",
        )
    except Exception as e:
        execution_time_ms = int((datetime.now() - start_time).total_seconds() * 1000)
        return ExecutionResult(
            executed=False,
            exit_code=-1,
            stdout="",
            stderr=str(e),
            execution_time_ms=execution_time_ms,
            output_files=[],
            error=str(e),
        )


def execute_locally(
    code: str,
    language: str,
    work_dir: Path,
    timeout: int = EXECUTION_TIMEOUT_SECONDS,
) -> ExecutionResult:
    """
    Execute code locally (fallback when Docker unavailable).

    WARNING: Less secure than Docker. Use only in trusted environments.

    Args:
        code: The code to execute
        language: Programming language
        work_dir: Working directory for execution
        timeout: Execution timeout in seconds

    Returns:
        ExecutionResult with exit code, output, and timing
    """
    start_time = datetime.now()

    # Write code to file
    if language == "python":
        script_file = work_dir / "script.py"
        script_file.write_text(code)
        cmd = ["python3", str(script_file)]
    elif language == "bash":
        script_file = work_dir / "script.sh"
        script_file.write_text(code)
        cmd = ["bash", str(script_file)]
    else:
        return ExecutionResult(
            executed=False,
            exit_code=-1,
            stdout="",
            stderr=f"Unsupported language: {language}",
            execution_time_ms=0,
            output_files=[],
            error=f"Unsupported language: {language}",
        )

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            timeout=timeout,
            text=True,
            cwd=work_dir,
        )

        execution_time_ms = int((datetime.now() - start_time).total_seconds() * 1000)

        # List output files
        output_files = [
            f.name for f in work_dir.iterdir()
            if f.is_file() and f.name not in ("script.py", "script.sh")
        ]

        return ExecutionResult(
            executed=True,
            exit_code=result.returncode,
            stdout=result.stdout,
            stderr=result.stderr,
            execution_time_ms=execution_time_ms,
            output_files=output_files,
        )

    except subprocess.TimeoutExpired:
        execution_time_ms = int((datetime.now() - start_time).total_seconds() * 1000)
        return ExecutionResult(
            executed=True,
            exit_code=-1,
            stdout="",
            stderr="Execution timed out",
            execution_time_ms=execution_time_ms,
            output_files=[],
            error="Timeout",
        )
    except Exception as e:
        execution_time_ms = int((datetime.now() - start_time).total_seconds() * 1000)
        return ExecutionResult(
            executed=False,
            exit_code=-1,
            stdout="",
            stderr=str(e),
            execution_time_ms=execution_time_ms,
            output_files=[],
            error=str(e),
        )


def verify_pdf_output(file_path: Path, expected: dict) -> tuple[bool, dict]:
    """
    Verify a PDF file meets expected criteria.

    Args:
        file_path: Path to the PDF file
        expected: Dict with expected properties (e.g., {"pages": 3})

    Returns:
        Tuple of (is_valid, actual_properties)
    """
    try:
        # Import here to avoid dependency if not needed
        from pypdf import PdfReader

        reader = PdfReader(file_path)
        page_count = len(reader.pages)

        properties = {"pages": page_count, "valid_format": True}

        # Check expected page count
        if "pages" in expected:
            if page_count != expected["pages"]:
                return False, properties

        # Check for expected text content
        if "text_contains" in expected:
            all_text = ""
            for page in reader.pages:
                all_text += page.extract_text() or ""

            for expected_text in expected["text_contains"]:
                if expected_text not in all_text:
                    properties["missing_text"] = expected_text
                    return False, properties

        return True, properties

    except ImportError:
        return False, {"error": "pypdf not installed (expected - neutral environment)"}
    except Exception as e:
        return False, {"error": str(e), "valid_format": False}


def verify_xlsx_output(file_path: Path, expected: dict) -> tuple[bool, dict]:
    """
    Verify an Excel file meets expected criteria.

    Args:
        file_path: Path to the XLSX file
        expected: Dict with expected properties

    Returns:
        Tuple of (is_valid, actual_properties)
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path)
        sheet_names = wb.sheetnames

        properties = {"sheets": sheet_names, "valid_format": True}

        # Check expected sheet
        if "sheet" in expected:
            if expected["sheet"] not in sheet_names:
                return False, properties

        return True, properties

    except ImportError:
        return False, {"error": "openpyxl not installed (expected - neutral environment)"}
    except Exception as e:
        return False, {"error": str(e), "valid_format": False}


def verify_json_output(file_path: Path, expected: dict) -> tuple[bool, dict]:
    """
    Verify a JSON file meets expected criteria.

    Args:
        file_path: Path to the JSON file
        expected: Dict with expected properties

    Returns:
        Tuple of (is_valid, actual_properties)
    """
    try:
        content = json.loads(file_path.read_text())
        properties = {"valid_json": True, "type": type(content).__name__}

        # Check expected keys
        if "keys" in expected and isinstance(content, dict):
            missing = [k for k in expected["keys"] if k not in content]
            if missing:
                properties["missing_keys"] = missing
                return False, properties

        return True, properties

    except json.JSONDecodeError as e:
        return False, {"error": str(e), "valid_json": False}
    except Exception as e:
        return False, {"error": str(e)}


def verify_yaml_output(file_path: Path, expected: dict) -> tuple[bool, dict]:
    """
    Verify a YAML file meets expected criteria.

    Args:
        file_path: Path to the YAML file
        expected: Dict with expected properties

    Returns:
        Tuple of (is_valid, actual_properties)
    """
    try:
        import yaml

        content = yaml.safe_load(file_path.read_text())
        properties = {"valid_yaml": True, "type": type(content).__name__}

        # Check expected keys
        if "keys" in expected and isinstance(content, dict):
            missing = [k for k in expected["keys"] if k not in content]
            if missing:
                properties["missing_keys"] = missing
                return False, properties

        return True, properties

    except ImportError:
        return False, {"error": "pyyaml not installed"}
    except Exception as e:
        return False, {"error": str(e), "valid_yaml": False}


def verify_code_syntax(file_path: Path, language: str) -> tuple[bool, dict]:
    """
    Verify code file has valid syntax.

    Args:
        file_path: Path to the code file
        language: Programming language

    Returns:
        Tuple of (is_valid, properties)
    """
    try:
        code = file_path.read_text()

        if language == "python":
            import ast
            ast.parse(code)
            return True, {"valid_syntax": True, "lines": code.count("\n") + 1}

        elif language == "json":
            json.loads(code)
            return True, {"valid_syntax": True}

        # For other languages, just check file exists and has content
        return True, {"valid_syntax": True, "size_bytes": len(code)}

    except SyntaxError as e:
        return False, {"valid_syntax": False, "error": str(e)}
    except Exception as e:
        return False, {"error": str(e)}


def verify_output(
    work_dir: Path,
    expected_files: list[str],
    expected_properties: dict,
) -> tuple[bool, dict]:
    """
    Verify execution produced expected outputs.

    Args:
        work_dir: Directory containing outputs
        expected_files: List of expected output filenames
        expected_properties: Dict of filename -> expected properties

    Returns:
        Tuple of (all_valid, file_properties)
    """
    all_valid = True
    properties = {}

    for filename in expected_files:
        file_path = work_dir / filename
        if not file_path.exists():
            properties[filename] = {"exists": False}
            all_valid = False
            continue

        properties[filename] = {"exists": True}
        expected = expected_properties.get(filename, {})

        # Verify based on file extension
        ext = file_path.suffix.lower()
        if ext == ".pdf":
            valid, props = verify_pdf_output(file_path, expected)
        elif ext == ".xlsx":
            valid, props = verify_xlsx_output(file_path, expected)
        elif ext == ".json":
            valid, props = verify_json_output(file_path, expected)
        elif ext in (".yaml", ".yml"):
            valid, props = verify_yaml_output(file_path, expected)
        elif ext == ".py":
            valid, props = verify_code_syntax(file_path, "python")
        else:
            # Generic file - just check it exists and has content
            content = file_path.read_bytes()
            valid = len(content) > 0
            props = {"size_bytes": len(content)}

        properties[filename].update(props)
        if not valid:
            all_valid = False

    return all_valid, properties


def verify_response(
    response: str,
    skill_name: str,
    prompt: str,
    expected_files: list[str],
    expected_properties: dict,
    use_docker: bool = True,
) -> VerificationResult:
    """
    Full verification pipeline: extract code, execute, verify outputs.

    Args:
        response: LLM response text
        skill_name: Name of the skill being evaluated
        prompt: Original prompt that generated the response
        expected_files: List of expected output filenames
        expected_properties: Dict of filename -> expected properties
        use_docker: Whether to use Docker (falls back to local if unavailable)

    Returns:
        VerificationResult with full verification details
    """
    # Step 1: Extract code
    extracted = extract_code(response)

    if not extracted.has_executable_code:
        return VerificationResult(
            skill_name=skill_name,
            prompt=prompt,
            code_extracted=False,
            code_language=extracted.primary_language,
            code_blocks_count=len(extracted.blocks),
            executed=False,
            execution_success=False,
            execution_error="No executable code found",
            execution_time_ms=0,
            output_files_created=[],
            output_valid=False,
        )

    # Step 2: Execute code
    with tempfile.TemporaryDirectory() as tmpdir:
        work_dir = Path(tmpdir)

        # Choose execution method
        if use_docker and is_docker_available():
            exec_result = execute_in_docker(
                extracted.combined_code,
                extracted.primary_language,
                work_dir,
            )
        else:
            exec_result = execute_locally(
                extracted.combined_code,
                extracted.primary_language,
                work_dir,
            )

        # Step 3: Verify outputs
        if exec_result.executed and exec_result.exit_code == 0:
            output_valid, output_properties = verify_output(
                work_dir,
                expected_files,
                expected_properties,
            )
        else:
            output_valid = False
            output_properties = {}

        return VerificationResult(
            skill_name=skill_name,
            prompt=prompt,
            code_extracted=True,
            code_language=extracted.primary_language,
            code_blocks_count=len(extracted.blocks),
            executed=exec_result.executed,
            execution_success=exec_result.exit_code == 0,
            execution_error=exec_result.error,
            execution_time_ms=exec_result.execution_time_ms,
            output_files_created=exec_result.output_files,
            output_valid=output_valid,
            output_properties=output_properties,
        )
